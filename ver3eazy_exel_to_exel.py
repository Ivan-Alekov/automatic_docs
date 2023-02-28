# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
import os, sys
import random, math

if not os.path.isdir('Отчет(ы)'):
    os.mkdir('Отчет(ы)')
os.chdir("Отчет(ы)")

path = "../template3.xlsx"
wb = openpyxl.reader.excel.load_workbook(filename=path, data_only=True)
wb.active = 0
sheet1 = wb.active

name_city = sheet1[f'B1'].value if sheet1[f'B1'].value != None else '-'
district = sheet1[f'B4'].value


def format_ws(ws, cell_range, f=13, st='thin'):
    font = Font(size=f, name='Times New Roman', color='000000')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(border_style=st, color='000000'),
                    right=Side(border_style=st, color='000000'),
                    top=Side(border_style=st, color='000000'),
                    bottom=Side(border_style=st, color='000000'))

    rows = [rows for rows in ws[cell_range]]
    flattened = [item for sublist in rows for item in sublist]
    [(setattr(cell, 'border', border), setattr(cell, 'font', font), setattr(cell, 'alignment', align)) for cell in
     flattened]


def format_ws1(ws, cell_range, f=13, st='thin'):
    font = Font(size=f, name='Times New Roman', color='000000')
    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side(border_style=st, color='000000'),
                    right=Side(border_style=st, color='000000'),
                    top=Side(border_style=st, color='000000'),
                    bottom=Side(border_style=st, color='000000'))

    rows = [rows for rows in ws[cell_range]]
    flattened = [item for sublist in rows for item in sublist]
    [(setattr(cell, 'border', border), setattr(cell, 'font', font), setattr(cell, 'alignment', align)) for cell in
     flattened]


def road(r, text1="", text2=""):
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    sheet.row_dimensions[r].height = 44.25
    sheet[f'A{r}'].value = f'{text1}'
    sheet[f'F{r}'].value = f'{text2}'
    format_ws1(sheet, cell_range=f'A{r}:F{r}')


def sheet_one(count_n=0, numb_k=0):
    """
    Работа с таблицей 'Общие данные' - ее расширение исходя из данных template
    :param count_n:
    :param numb_k:
    :return:
    """
    type_of_covering = ['капитальный', 'облегченный', 'переходный', 'низший']
    row = 8

    sheet[f'F5'].value = go_word['name_road']
    sheet[f'F6'].value = go_word['year']
    #  нахождение всех длин участков дорог
    lenght_all = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k}'].value,
                                                                   sheet1[f'T{numb_k}'].value,
                                                                   sheet1[f'U{numb_k}'].value,
                                                                   sheet1[f'V{numb_k}'].value]))
    l_all = [(ind, znach) for ind, znach in enumerate(lenght_all) if znach != 0]
    # нахождение всех ширин участков дорог
    weight_all = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'J{numb_k}'].value,
                                                                   sheet1[f'K{numb_k}'].value,
                                                                   sheet1[f'L{numb_k}'].value,
                                                                   sheet1[f'M{numb_k}'].value,
                                                                   sheet1[f'N{numb_k}'].value,
                                                                   sheet1[f'O{numb_k}'].value]))
    w_all = [(ind, znach) for ind, znach in enumerate(weight_all) if znach != 0]

    q = count_n

    #  заполнение пункта: Адрес(а) участка(ов) автомобильной дороги (улицы)
    sheet[f'F7'].value = f'0+000-{go_word["covering_weight1"]}+{go_word["covering_weight2"]}' if q < 2 else ''
    format_ws1(sheet, cell_range=f'A{7}:F{7}')

    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            sheet[f'F{row + i}'].value = f'{sheet1[f"R{numb_k + i}"].value}'
        row += q
        format_ws(sheet, cell_range=f'A8:F{row - 1}')

    #  заполнение пункта: Протяженность автомобильной дороги (улицы) или ее участка(ов) (м)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    sheet.row_dimensions[row].height = 44.25

    sheet[f'A{row}'].value = f'Протяженность автомобильной дороги (улицы) или ее участка(ов) (м):'
    len_1 = f'{float(go_word["lenght_all"].replace(",", ".")):.3f}'.split(".")
    sheet[f'F{row}'].value = (f'{len_1[0] if len_1[0] != "0" else ""}{len_1[1]}') if q < 2 else ''
    format_ws1(sheet, cell_range=f'A{row}:F{row}')

    row += 1
    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            lenght_all1 = list(map(lambda x: x if x != None else 0, [sheet1[f'S{numb_k + i}'].value,
                                                                           sheet1[f'T{numb_k + i}'].value,
                                                                           sheet1[f'U{numb_k + i}'].value,
                                                                           sheet1[f'V{numb_k + i}'].value]))
            l_all1 = [(ind, znach) for ind, znach in enumerate(lenght_all1) if znach != 0]
            len_1 = f'{l_all1[0][1]:.3f}'.split(".")
            sheet[f'F{row + i}'].value = f'{len_1[0] if len_1[0] != "0" else ""}{len_1[1]}'
        row += q
        format_ws(sheet, cell_range=f'A{row - q}:F{row - 1}')
    #  заполнение пункта: ТЕХНИЧЕСКИЕ ХАРАКТЕРИСТИКИ
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sheet.row_dimensions[row].height = 44.25
    sheet[f'A{row}'].value = f'2. ТЕХНИЧЕСКИЕ ХАРАКТЕРИСТИКИ'
    format_ws(sheet, cell_range=f'A{row}:F{row}')
    row += 1

    #  заполнение пункта: Категория автомобильной дороги (улицы) в соответствии с СП 42.13330.2016
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    sheet.row_dimensions[row].height = 44.25
    sheet[f'A{row}'].value = f'Категория автомобильной дороги (улицы) в соответствии с СП 42.13330.2016:'
    sheet[f'F{row}'].value = f'{sheet1[f"W{numb_k}"].value.lower()}' if q < 2 else ''
    format_ws1(sheet, cell_range=f'A{row}:F{row}')

    row += 1
    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            sheet[f'F{row + i}'].value = f'{sheet1[f"W{numb_k + i}"].value.lower()}'
        row += q
        format_ws(sheet, cell_range=f'A{row - q}:F{row - 1}')

    #  заполнение пункта: Группа дороги (улицы) по ГОСТ Р 50597-2017
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    sheet.row_dimensions[row].height = 44.25
    sheet[f'A{row}'].value = f'Группа дороги (улицы) по ГОСТ Р 50597-2017:'
    sheet[f'F{row}'].value = f'{sheet1[f"AC{numb_k}"].value}'.upper() if q < 2 else ''
    format_ws1(sheet, cell_range=f'A{row}:F{row}')
    row += 1
    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            sheet[f'F{row + i}'].value = f'{sheet1[f"AC{numb_k + i}"].value}'.upper()
        row += q
        format_ws(sheet, cell_range=f'A{row - q}:F{row - 1}')

    #  заполнение пункта: Число полос движения
    road(row, f'Число полос движения:', f'' if q > 1 else 1 if w_all[0][1] < 4.5 else 2)

    row += 1
    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            weight_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'J{numb_k + i}'].value,
                                                                           sheet1[f'K{numb_k + i}'].value,
                                                                           sheet1[f'L{numb_k + i}'].value,
                                                                           sheet1[f'M{numb_k + i}'].value,
                                                                           sheet1[f'N{numb_k + i}'].value,
                                                                           sheet1[f'O{numb_k + i}'].value]))
            w_all1 = [(ind, znach) for ind, znach in enumerate(weight_all1) if znach != 0]
            sheet[f'F{row + i}'].value = f'{1 if w_all1[0][1] < 4.5 else 2}'
        row += q
        format_ws(sheet, cell_range=f'A{row - q}:F{row - 1}')

    #  заполнение пункта: Тип дорожной одежды
    type = type_of_covering[l_all[0][0]]
    road(row, f'Тип дорожной одежды:',f'{type}' if q < 2 else '')

    row += 1
    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            lenght_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k + i}'].value,
                                                                           sheet1[f'T{numb_k + i}'].value,
                                                                           sheet1[f'U{numb_k + i}'].value,
                                                                           sheet1[f'V{numb_k + i}'].value]))
            l_all1 = [(ind, znach) for ind, znach in enumerate(lenght_all1) if znach != 0]
            type = type_of_covering[l_all1[0][0]]
            sheet[f'F{row + i}'].value = f'{type}'.lower()
        row += q
        format_ws(sheet, cell_range=f'A{row - q}:F{row - 1}')

    #  заполнение пункта: Ширина покрытия (м)
    road(row, f'Ширина покрытия (м):', f'' if q > 1 else str(w_all[0][1]).replace('.', ","))
    row += 1
    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            weight_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'J{numb_k + i}'].value,
                                                                            sheet1[f'K{numb_k + i}'].value,
                                                                            sheet1[f'L{numb_k + i}'].value,
                                                                            sheet1[f'M{numb_k + i}'].value,
                                                                            sheet1[f'N{numb_k + i}'].value,
                                                                            sheet1[f'O{numb_k + i}'].value]))
            w_all1 = [(ind, znach) for ind, znach in enumerate(weight_all1) if znach != 0]
            sheet[f'F{row + i}'].value = f'{w_all1[0][1]}'.replace(".", ",")
        row += q
        format_ws(sheet, cell_range=f'A{row - q}:F{row - 1}')

    #  заполнение пункта: Интенсивность движения (авт./сут)
    road(row, f'Интенсивность движения (авт./сут):', f'' if q > 1 else f'{sheet1[f"X{numb_k}"].value}')
    row += 1
    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            sheet[f'F{row + i}'].value = f'{sheet1[f"X{numb_k + i}"].value}'
        row += q
        format_ws(sheet, cell_range=f'A{row - q}:F{row - 1}')

    #  заполнение пункта: Расчетная скорость движения
    speed_30 = ['проезд', 'местная дорога']
    speed_40 = ['местная улица', "основная улица"]
    road(row, f'Расчетная скорость движения:', f'' if q > 1 else '30 км/ч' if sheet1[f"W{numb_k}"].value.lower() in speed_30 else '40 км/ч')
    row += 1
    if q > 1:
        for i in range(q):
            sheet.insert_rows(row + i, amount=1)
            sheet.row_dimensions[row + i].height = 44.25
            sheet.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=5)
            sheet[f'A{row + i}'].value = f'Участок №{i + 1}'
            sheet[f'F{row + i}'].value = '30 км/ч' if sheet1[f"W{numb_k + i}"].value.lower() in speed_30 else '40 км/ч'
        row += q
        format_ws(sheet, cell_range=f'A{row - q}:F{row - 1}')
    sheet.print_area = f"A1:F{row - 1}"


def sheet_two(count_n=0, numb_k=0):
    """
    Работа с таблицей: Состояние ДО
    :param count_n:
    :param numb_k:
    :return:
    """
    def for_sheet_2(r=8, count_a=0, numb_b=100, stop=0, n=0):
        sheet[f'A{r}'].value = f'{int(math.ceil(count_a * 1000)) if stop==1 else int(math.ceil(count_a))}'
        sheet[f'B{r}'].value = f'{int(math.ceil(numb_b))}'
        c = n + random.randint(1, 9) / 10
        while c == sheet[f'C{r - 1}'].value:
            c = n + random.randint(1, 9) / 10
        sheet[f'C{r}'].value = f'{c:.1f}'.replace(".", ",")
        sheet[f'D{r}'].value = f'{c:.1f}'.replace(".", ",")
        sheet[f'E{r}'].value = f'4'
        sheet[f'F{r}'].value = f'{"норматив" if n == 4 else "ненорматив"}'
        sheet[f'G{r}'].value = f'{"норматив" if n == 4 else "ненорматив"}'

    def loop_for_2(r=0, max_l=0, norm1=0):
        lenght = 100
        big_len = 100
        div_l, mod_l = divmod(max_l * 1000, 100)
        for i in range(1, int(div_l) + 1):
            sheet.insert_rows(r + i - 1, amount=1)
            for_sheet_2(r=r + i - 1, count_a=big_len * i, numb_b=lenght, n=norm1)
        r += int(div_l)
        if mod_l != 0:
            if (mod_l <= 20 and sheet[f'A{r - 1}'].value == 100) or (
                    mod_l <= 20 and big_len >= 100):
                for_sheet_2(r=r - 1, count_a=max_l, numb_b=lenght + mod_l, n=norm1, stop=1)
            else:
                sheet.insert_rows(r, amount=1)
                for_sheet_2(r=r, count_a=max_l, numb_b=mod_l, n=norm1, stop=1)
                r += 1
        return r

    lenght_all = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k}'].value,
                                                                   sheet1[f'T{numb_k}'].value,
                                                                   sheet1[f'U{numb_k}'].value,
                                                                   sheet1[f'V{numb_k}'].value]))
    l_all = [(ind, znach) for ind, znach in enumerate(lenght_all) if znach != 0]

    row = 8
    q = count_n
    res_match_yes = []
    res_match_not = []

    if q < 2:
        all = l_all[0][1]
        if sheet1[f'Q{numb_k}'].value.lower() == 'соотв.':
            norm = 4.0
        else:
            norm = 3.0
        row = loop_for_2(r=row, max_l=all, norm1=norm)
    else:
        for i in range(q):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            sheet[f'A{row}'].value = f'Участок №{i + 1}'
            sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row += 1
            lenght_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k + i}'].value,
                                                                           sheet1[f'T{numb_k + i}'].value,
                                                                           sheet1[f'U{numb_k + i}'].value,
                                                                           sheet1[f'V{numb_k + i}'].value]))
            l_all1 = [(ind, znach) for ind, znach in enumerate(lenght_all1) if znach != 0]
            if sheet1[f'Q{numb_k + i}'].value.lower() == 'соотв.':
                norm = 4.0
                res_match_yes.append(str(i + 1))
            else:
                norm = 3.0
                res_match_not.append(str(i + 1))
            row = loop_for_2(r=row, max_l=l_all1[0][1], norm1=norm)
    row = row - 1
    format_ws(sheet, cell_range=f'A{8}:G{row}', f=11)

    #  Добавление подсчета подходящих покрытий в %
    sheet.merge_cells(start_row=row + 2, start_column=2, end_row=row + 2, end_column=7)
    sheet[f'B{row + 2}'].value = '-Нормативное состояние покрытия в %'
    sheet[f'B{row + 2}'].font = Font(size=11, name='Times New Roman', color='000000')
    sheet[f'B{row + 2}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    sheet.row_dimensions[row + 3].height = 47.25
    sheet.merge_cells(start_row=row + 3, start_column=1, end_row=row + 3, end_column=7)

    res_sort_yes = ", ".join(res_match_yes)
    res_sort_yes = "на участке (ах) №" + res_sort_yes + " нормативное" if res_sort_yes else ""
    res_sort_not = ", ".join(res_match_not)
    res_sort_not = "на участке (ах) №" + res_sort_not + " ненормативное" if res_sort_not else ""
    text = ', '.join(i for i in [res_sort_yes, res_sort_not] if len(i) != 0)
    if len(text.split(',')) == 1:
        text = 'нормативное' if sheet1[f'Q{numb_k}'].value.lower() == 'соотв.' else "ненормативное"

    conclusion = f'Заключение: Дорожная одежда (покрытие) автомобильной дороги (улицы), с ' \
                                 f'учетом бальной оценки выполненной в соответствии с таблицей 4.5 ОДМ ' \
                                 f'218.4.039-2018 имеет {text} состояние.'
    result[0] = conclusion
    sheet[f'A{row + 3}'].value = conclusion
    sheet[f'A{row + 3}'].font = Font(size=11, name='Times New Roman', color='000000')
    sheet[f'A{row + 3}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    sheet.print_area = f"A1:G{row + 3}"

    return row + 3


def sheet_three(count_n=0, numb_k=0):
    """
    Работа с таблицей: Состояние ЗП
    :param count_n:
    :param numb_k:
    :return:
    """
    lenght_all = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k}'].value,
                                                                   sheet1[f'T{numb_k}'].value,
                                                                   sheet1[f'U{numb_k}'].value,
                                                                   sheet1[f'V{numb_k}'].value]))
    l_all = [(ind, znach) for ind, znach in enumerate(lenght_all) if znach != 0]

    q = count_n
    cell = 2
    row = 2
    res_match_yes = []
    res_match_not = []
    if q > 1:
        sheet.cell(2, cell + q - 1).value = f'Таблица 2'
        sheet.cell(2, cell + q - 1).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        sheet.cell(2, cell + q - 1).font = Font(size=12, name='Times New Roman', color='000000')

        sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=cell + q - 1)
        sheet.cell(3, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=cell + q - 1)
        sheet.cell(6, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        width = 49.71 + 17.71 * 2
        for i in range(q):
            lenght_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k + i}'].value,
                                                                           sheet1[f'T{numb_k + i}'].value,
                                                                           sheet1[f'U{numb_k + i}'].value,
                                                                           sheet1[f'V{numb_k + i}'].value]))
            l_all1 = [(ind, znach) for ind, znach in enumerate(lenght_all1) if znach != 0]
            len1 = f'{l_all1[0][1]:.3f}'.split('.')
            sheet.column_dimensions[chr(ord("A") + cell + i - 1)].width = width // q
            sheet.cell(4, cell + i).value = f'Участок №{i + 1}'

            sheet.cell(5, cell + i).value = f'0+{len1[0]}' if len1[0] == 0 else f'{len1[0]}+{len1[1]}'

            sheet.cell(7, cell + i).value = sheet1[f"Y{numb_k + i}"].value

            sheet.cell(8, cell + i).value = sheet1[f"Z{numb_k + i}"].value + " НТД"
            if sheet1[f"Y{numb_k + i}"].value.lower() == 'соотв.':
                res_match_yes.append(str(i + 1))
            else:
                res_match_not.append(str(i + 1))
    else:
        sheet.column_dimensions[chr(ord("A"))].width = 49.71
        sheet.cell(2, cell).value = f'Таблица 2'
        sheet.cell(2, cell).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        sheet.cell(2, cell).font = Font(size=12, name='Times New Roman', color='000000')

        sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=cell)
        sheet.cell(3, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sheet.cell(6, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sheet.cell(4, cell).value = f'Адрес участка, м'

        len1 = f'{l_all[0][1]:.3f}'.split('.')
        sheet.cell(5, cell).value = f'0+000-0+{len1[0]}' if len1[0] == 0 else f'0+000-{len1[0]}+{len1[1]}'

        sheet.cell(7, cell).value = sheet1[f"Y{numb_k}"].value

        sheet.cell(8, cell).value = sheet1[f"Z{numb_k}"].value + " НТД"

    format_ws(sheet,cell_range=f"A4:{chr((ord('A') + q))}8", f=11)

    sheet.row_dimensions[44].height = 24
    sheet.merge_cells(start_row=44, start_column=1, end_row=44, end_column=cell + q - 1)

    res_sort_yes = ", ".join(res_match_yes)
    res_sort_yes = "на участке(ах) №" + res_sort_yes + " соответствие" if res_sort_yes else ""
    res_sort_not = ", ".join(res_match_not)
    res_sort_not = "на участке(ах) №" + res_sort_not + " несоответствие" if res_sort_not else ""
    text = ', '.join(i for i in [res_sort_yes, res_sort_not] if len(i) != 0)
    if len(text.split(',')) == 1:
        text = 'соответствие' if sheet1[f'Y{numb_k}'].value.lower() == 'соотв.' else "несоответствие"
    conclusion = f'Заключение: При визуальной оценке состояния земляного полотна и ' \
                 f'водоотвода установлено {text} нормативным требованиям.'

    result[1] = conclusion
    sheet.row_dimensions[46].height = 64.5
    sheet.merge_cells(start_row=46, start_column=1, end_row=46, end_column=cell + q - 1)
    sheet[f'A46'].value = conclusion

    sheet.print_area = f"A1:{chr((ord('A') + q))}46"

    return 46


def sheet_four(count_n=0, numb_k=0):
    """
    Работа с таблицей: Ширина покрытия
    :param count_n:
    :param numb_k:
    :return:
    """
    def for_sheet_4(r=8, count=0, numb=0, stop=0, w_max=0):
        sheet[f'A{r}'].value = f'{numb}'
        sheet[f'B{r}'].value = f'{int(math.ceil(count))}'
        sheet[f'C{r}'].value = f'{w_max + (random.randint(-9, -1) / 100 if random.randint(0, 1) == 0 else random.randint(1, 9) / 100):.2f}'.replace(".", ",")

    def loop_for_4(r, stop, max_w=0):
        lenght = 0
        big_len = 0
        i = 0
        while big_len * 1000 + lenght <= stop:
            if lenght % 1000 == 0 and lenght != 0:
                lenght = 0
                big_len += 1
                sheet.insert_rows(r + i, amount=1)
                for_sheet_4(r + i, count=lenght, numb=big_len, w_max=max_w)
            else:
                sheet.insert_rows(r + i, amount=1)
                for_sheet_4(r + i, count=lenght, numb=big_len, w_max=max_w)
            lenght += 100
            i += 1
        r += i

        if stop % 100 != 0:
            if (stop % 100 <= 20 and sheet[f'B{r - 1}'].value == 0 and sheet[f'A{r - 1}'].value != 0) or (stop % 100 <= 20 and lenght >= 100 and r != 9):
                for_sheet_4(r=r - 1, count=stop % 100 + lenght - 100 if lenght != 0 else stop % 100, numb=big_len,
                            w_max=max_w)
            else:
                sheet.insert_rows(r, amount=1)
                for_sheet_4(r=r, count=stop % 100 + lenght - 100 if lenght != 0 else stop % 100, numb=big_len, w_max=max_w)
                r += 1
        return r

    lenght_all = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k}'].value,
                                                                   sheet1[f'T{numb_k}'].value,
                                                                   sheet1[f'U{numb_k}'].value,
                                                                   sheet1[f'V{numb_k}'].value]))
    l_all = [(ind, znach) for ind, znach in enumerate(lenght_all) if znach != 0]
    weight_all = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'J{numb_k}'].value,
                                                                   sheet1[f'K{numb_k}'].value,
                                                                   sheet1[f'L{numb_k}'].value,
                                                                   sheet1[f'M{numb_k}'].value,
                                                                   sheet1[f'N{numb_k}'].value,
                                                                   sheet1[f'O{numb_k}'].value]))
    w_all = [(ind, znach) for ind, znach in enumerate(weight_all) if znach != 0]

    row = 8
    q = count_n
    count_45 = []
    count_46 = []

    if q < 2:
        all = float(go_word['lenght_all'].replace(",", ".")) * 1000
        row = loop_for_4(row, all,max_w=w_all[0][1])
    else:
        for n in range(q):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            sheet[f'A{row}'].value = f'Участок №{n + 1}'
            sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row += 1
            lenght_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k + n}'].value,
                                                                           sheet1[f'T{numb_k + n}'].value,
                                                                           sheet1[f'U{numb_k + n}'].value,
                                                                           sheet1[f'V{numb_k + n}'].value]))
            l_all1 = [(ind, znach) for ind, znach in enumerate(lenght_all1) if znach != 0]
            weight_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'J{numb_k + n}'].value,
                                                                           sheet1[f'K{numb_k + n}'].value,
                                                                           sheet1[f'L{numb_k + n}'].value,
                                                                           sheet1[f'M{numb_k + n}'].value,
                                                                           sheet1[f'N{numb_k + n}'].value,
                                                                           sheet1[f'O{numb_k + n}'].value]))
            w_all1 = [(ind, znach) for ind, znach in enumerate(weight_all1) if znach != 0]
            row = loop_for_4(row, float(l_all1[0][1]) * 1000, max_w=w_all1[0][1])
            if sheet1[f"AA{numb_k + n}"].value.lower() == 'соотв.':
                count_45.append(str(n + 1))
            else:
                count_46.append(str(n + 1))
    row = row - 1
    format_ws(sheet, cell_range=f'A{8}:C{row}', f=12)


    sheet.row_dimensions[row + 2].height = 15.75
    sheet.merge_cells(start_row=row + 2, start_column=1, end_row=row + 8, end_column=3)

    res_sort_weight_45 =", ".join(count_45)
    res_sort_weight_45 = "участке(ах) №" + res_sort_weight_45 + " соответствует" if res_sort_weight_45 else ""
    res_sort_weight_46 = ", ".join(count_46)
    res_sort_weight_46 = "участке(ах) №" + res_sort_weight_46 + " несоответствует" if res_sort_weight_46 else ""
    text = ', '.join(i for i in [res_sort_weight_45, res_sort_weight_46] if len(i) != 0)
    if len(text.split(',')) == 1:
        text = 'соответствует' if sheet1[f"AA{numb_k}"].value.lower() == 'соотв.' else "несоответствует"
    conclusion = f'Заключение: ширина полос движения (покрытия) {text} требованиям таблицы ' \
                                 f'11.4 СП 42.13330.2016 "Градостроительство. Планировка и застройка городских и ' \
                                 f'сельских поселений", для установленной категории автомобильной дороги (улицы), ' \
                                 f'так как {"" if res_sort_weight_46 else "не"} превышает допустимое отклонение 0,5 м установленное п.5.2.2.1 ОДМ 218.4.039-2018 ' \
                                 f'от требуемой ширины, установленной для категории улицы в соответствии с требованиями СП 42.13330.2016'
    result[2] = conclusion
    sheet[f'A{row + 2}'].value = conclusion
    sheet[f'A{row + 2}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet.print_area = f"A1:C{row + 8}"

    return row + 2


def sheet_five(count_n=0, numb_k=0):
    """
    Работа с таблицей: Продольные уклоны
    :param count_n:
    :param numb_k:
    :return:
    """
    def for_sheet_5(r=8, count=0, numb=0, stop=0, w_max=0):
        sheet[f'A{r}'].value = f'{numb}'
        sheet[f'B{r}'].value = f'{int(math.ceil(count))}'
        sheet[f'C{r}'].value = f'{float(random.randint(4, 20))}'.replace(".", ",")

    def loop_for_5(r, stop, max_w=0):
        lenght = 0
        big_len = 0
        i = 0
        while big_len * 1000 + lenght <= stop:
            if lenght % 1000 == 0 and lenght != 0:
                lenght = 0
                big_len += 1
                sheet.insert_rows(r + i, amount=1)
                for_sheet_5(r + i, count=lenght, numb=big_len, w_max=max_w)
            else:
                sheet.insert_rows(r + i, amount=1)
                for_sheet_5(r + i, count=lenght, numb=big_len, w_max=max_w)
            lenght += 100
            i += 1
        r += i
        if stop % 100 != 0:
            if (stop % 100 <= 20 and sheet[f'B{r - 1}'].value == 0 and sheet[f'A{r - 1}'].value != 0) or (stop % 100 <= 20 and lenght >= 100 and r != 9):
                for_sheet_5(r=r - 1, count=stop % 100 + lenght - 100 if lenght != 0 else stop % 100, numb=big_len,
                            w_max=max_w)
            else:
                sheet.insert_rows(r, amount=1)
                for_sheet_5(r=r, count=stop % 100 + lenght - 100 if lenght != 0 else stop % 100, numb=big_len,
                            w_max=max_w)
                r += 1
        return r

    lenght_all = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k}'].value,
                                                                   sheet1[f'T{numb_k}'].value,
                                                                   sheet1[f'U{numb_k}'].value,
                                                                   sheet1[f'V{numb_k}'].value]))
    l_all = [(ind, znach) for ind, znach in enumerate(lenght_all) if znach != 0]
    weight_all = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'J{numb_k}'].value,
                                                                   sheet1[f'K{numb_k}'].value,
                                                                   sheet1[f'L{numb_k}'].value,
                                                                   sheet1[f'M{numb_k}'].value,
                                                                   sheet1[f'N{numb_k}'].value,
                                                                   sheet1[f'O{numb_k}'].value]))
    w_all = [(ind, znach) for ind, znach in enumerate(weight_all) if znach != 0]

    row = 8
    q = count_n

    if q < 2:
        all = float(go_word['lenght_all'].replace(",", ".")) * 1000
        row = loop_for_5(row, all, max_w=w_all[0][1])
    else:
        for n in range(q):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            sheet[f'A{row}'].value = f'Участок №{n + 1}'
            sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row += 1
            lenght_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'S{numb_k + n}'].value,
                                                                           sheet1[f'T{numb_k + n}'].value,
                                                                           sheet1[f'U{numb_k + n}'].value,
                                                                           sheet1[f'V{numb_k + n}'].value]))
            l_all1 = [(ind, znach) for ind, znach in enumerate(lenght_all1) if znach != 0]
            weight_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'J{numb_k + n}'].value,
                                                                           sheet1[f'K{numb_k + n}'].value,
                                                                           sheet1[f'L{numb_k + n}'].value,
                                                                           sheet1[f'M{numb_k + n}'].value,
                                                                           sheet1[f'N{numb_k + n}'].value,
                                                                           sheet1[f'O{numb_k + n}'].value]))
            w_all1 = [(ind, znach) for ind, znach in enumerate(weight_all1) if znach != 0]
            row = loop_for_5(row, float(l_all1[0][1]) * 1000, max_w=w_all1[0][1])
    row = row - 1
    format_ws(sheet, cell_range=f'A{8}:C{row}', f=12)

    sheet.row_dimensions[row + 2].height = 15.75
    sheet.merge_cells(start_row=row + 2, start_column=1, end_row=row + 10, end_column=3)

    conclusion = f'Заключение: Продольные уклоны автомобильной дороги (улицы) соответствует требованиям ' \
                                 f'таблицы 11.4 СП 42.13330.2016 "Градостроительство. Планировка и застройка городских и ' \
                                 f'сельских поселений", для установленной категории автомобильной дороги (улицы), ' \
                                 f'так как не превышают допустимое отклонение 20% установленное п.5.2.2.5 ОДМ 218.4.039-2018 ' \
                                 f'от требуемых значений установленных для категории улицы в соответствии с требованиями СП 42.13330.2016'
    result[3] = conclusion
    sheet[f'A{row + 2}'].value = conclusion
    sheet[f'A{row + 2}'].font = Font(size=10, name='Times New Roman', color='000000')
    sheet[f'A{row + 2}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet.print_area = f"A1:C{row + 10}"
    return row + 2


def sheet_six(count_n=0, numb_k=0):
    """
    Работа с таблицей: Обустройство
    :param count_n:
    :param numb_k:
    :return:
    """
    q = count_n
    row = 7
    if q > 1:
        for i in range(q):
            sheet.row_dimensions[row + 2].height = 30.75
            sheet[f'A{row}'].value = f'Участок №{i + 1}'
            sheet[f'B{row}'].value = f'Дорожные знаки утрачены'
            sheet[f'B{row + 1}'].value = f'Искусственное освещение отсутствует'
            sheet[f'B{row + 2}'].value = f'Отсутствуют предусмотренные СП 42.13330.2016 тротуары или пешеходные дорожки'
            sheet.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=1)
            row += 3
        sheet.print_area = f"A1:B{row - 1}"
    else:
        sheet.merge_cells(start_row=7, start_column=1, end_row=9, end_column=1)
        sheet[f'A7'].value = f'от км 0+000 до км {go_word["covering_weight1"]}+{go_word["covering_weight2"]}'
        sheet.print_area = f"A1:B9"
    format_ws(sheet, cell_range=f"A7:A{row}",f=12)
    format_ws1(sheet, cell_range=f"B7:B{row}", f=12)


def sheet_seven(ans=False):
    """
    Работа с таблицей: Заключение
    :param ans:
    :return:
    """
    def add_text(count_size=0, end_r=0, text_b="", text_a=""):
        sheet[f'B{count_size}'].value = text_b
        sheet[f'A{count_size}'].value = text_a
        sheet.merge_cells(start_row=count_size, start_column=2, end_row=count_size + end_r, end_column=2)
        sheet.merge_cells(start_row=count_size, start_column=1, end_row=count_size + end_r, end_column=1)
        return count_size + end_r + 1

    sheet[f'B5'].value = go_word['name_road']
    text_in = f'Требуется проведение работ по уширению проезжей части, восстановлению покрытия, устройству недостающих ' \
              f'тротуаров или пешеходных дорожек, устройству недостающих элементов освещения, устройству недостающих ' \
              f'дорожных знаков. В соответствии с классификацией работ по капитальному ремонту, ремонту и содержанию ' \
              f'автомобильных дорог утвержденной  приказом министерства транспорта Российской Федерации от 16 ноября ' \
              f'2012 года N 402, и СП 34.13330.2012 г. выполнение указанных мероприятий возможно в рамках проведения ' \
              f'работ по капитальному ремонту автомобильной дороги (улицы), а в случае необходимости дополнительного ' \
              f'отвода земельных участков и (или) повышения категории автомобильной дороги (улицы), требуется ' \
              f'проведение работ по реконструкции автомобильной дороги (улицы).'

    numb = 70 if ans else 47
    for i in range(numb):
        sheet.insert_rows(i + 8)
        sheet.row_dimensions[i + 8].height = 15.75
    count_size = 8
    text = 'Оценка состояния дорожной одежды (покрытия)'
    count_size = add_text(count_size=count_size, end_r=5, text_b=result[0], text_a=text)

    text = 'Оценка состояния земляного полотна и водоотвода'
    count_size = add_text(count_size=count_size, end_r=5, text_b=result[1], text_a=text)

    text = 'Результаты оценки по параметру "ширина полосы движения" (покрытия)'
    count_size = add_text(count_size=count_size, end_r=7, text_b=result[2], text_a=text)

    text = 'Оценка автомобильной дороги (улицы) по параметру "продольный уклон"'
    count_size = add_text(count_size=count_size, end_r=7, text_b=result[3], text_a=text)

    sheet[f'B{count_size}'].value = 'Дорожные знаки отсутствуют'
    sheet[f'B{count_size + 1}'].value = "Сигнальные столбики отсутствуют"
    sheet[f'A{count_size}'].value = 'Оценка  состояния элементов инженерного обустройства'
    sheet.merge_cells(start_row=count_size, start_column=1, end_row=count_size + 1, end_column=1)
    format_ws1(sheet, cell_range=f'A8:B22', f=12)
    count_size += 2

    if ans:
        text = 'Результаты оценки по параметру "продольная ровность"'
        count_size = add_text(count_size=count_size, end_r=7, text_a=text)

        text = 'Результаты оценки по параметру "поперечная ровность"'
        count_size = add_text(count_size=count_size, end_r=7, text_a=text)

        text = 'Результаты оценки по параметру "коэффициент сцепления колеса автомобиля с дорожным покрытием"'
        count_size = add_text(count_size=count_size, end_r=7, text_a=text)

        count_size += 3
        sheet[f'A{count_size}'].value = text_in
        sheet.merge_cells(start_row=count_size, start_column=1, end_row=count_size + 7, end_column=2)
        sheet[f'A{count_size}'].font = Font(size=13, name='Times New Roman', color='000000')
        sheet[f'A{count_size}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        count_size += 10
        sheet[f'A{count_size}'].value = 'Руководитель группы диагностики а/д     ________   И.В. Дацко'
        sheet.merge_cells(start_row=count_size, start_column=1, end_row=count_size, end_column=2)
        sheet[f'A{count_size}'].font = Font(size=15, name='Times New Roman', color='000000', bold=True)
        sheet[f'A{count_size}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    if not(ans):
        count_size += 3
        sheet[f'A{count_size}'].value = text_in
        sheet.merge_cells(start_row=count_size, start_column=1, end_row=count_size + 7, end_column=2)
        sheet[f'A{count_size}'].font = Font(size=13, name='Times New Roman', color='000000')
        sheet[f'A{count_size}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        count_size += 10
        sheet[f'A{count_size}'].value = 'Руководитель группы диагностики а/д     ________   И.В. Дацко'
        sheet.merge_cells(start_row=count_size, start_column=1, end_row=count_size, end_column=2)
        sheet[f'A{count_size}'].font = Font(size=15, name='Times New Roman', color='000000', bold=True)
        sheet[f'A{count_size}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    if ans is False:
        rmv = wb1['Продольная ровность_acf']
        wb1.remove(rmv)

        rmv = wb1['поперечная ровность (колея)_acf']
        wb1.remove(rmv)

        rmv = wb1['коэффициент сцепления (2)_acf']
        wb1.remove(rmv)
    minus = 10 if not ans else 9
    format_ws1(sheet, cell_range=f'B{8}:B{numb - minus}', f=12)
    format_ws(sheet, cell_range=f'A{8}:B{numb - minus}', f=12)
    sheet.print_area = f"A1:B{numb + 5}" if not (ans) else f"A1:B{numb + 6}"
    #print(result[1], sheet[f'B{13}'].value, result[1] == sheet[f'B{13}'].value, sep='\n')


k = 10
count_name_road = 0
baaaaad = set()

while True:
    result = {}
    if sheet1[f'A{k}'].value == "КОНЕЦ":
        break
    path1 = "../example.xlsx"
    wb1 = openpyxl.reader.excel.load_workbook(filename=path1, data_only=True)
    count_none = 0
    if sheet1[f'A{k}'].value == None:
        count_name_road += 1
        k += 1
        continue
    else:
        count_name_road = 0
        while True:
            if sheet1[f'A{k + count_none + 1}'].value == None:
                count_none += 1
            else:
                break
    for_name_road = sheet1[f'A{k - count_name_road}'].value
    for_year = sheet1[f'B{k - count_name_road}'].value
    for_length = sheet1[f'C{k - count_name_road}'].value

    result_read_exel = [for_name_road, for_year, for_length,
                        sheet1[f'D{k}'].value, sheet1[f'E{k}'].value, sheet1[f'F{k}'].value,
                        sheet1[f'G{k}'].value, sheet1[f'H{k}'].value, sheet1[f'I{k}'].value,
                        sheet1[f'J{k}'].value, sheet1[f'K{k}'].value, sheet1[f'L{k}'].value,
                        sheet1[f'M{k}'].value, sheet1[f'N{k}'].value, sheet1[f'O{k}'].value,
                        sheet1[f'P{k}'].value,
                        sheet1[f'Q{k}'].value, sheet1[f'R{k}'].value, sheet1[f'S{k}'].value,
                        sheet1[f'T{k}'].value, sheet1[f'U{k}'].value, sheet1[f'V{k}'].value,
                        sheet1[f'W{k}'].value, sheet1[f'X{k}'].value ,sheet1[f'Y{k}'].value,
                        sheet1[f'Z{k}'].value, sheet1[f'AA{k}'].value, sheet1[f'AB{k}'].value,
                        sheet1[f'AC{k}'].value]

    go_word = {'name_road': result_read_exel[0],
               'year': int(result_read_exel[1]) if result_read_exel[1] != None else 'данные отсутствуют',
               'lenght_all': f'{result_read_exel[2]:.3f}'.replace('.', ','),
               'name_city': name_city, 'covering_weight': [],
               'asphalt_l': f'{result_read_exel[3]:.3f}'.replace(".", ',') if result_read_exel[3] != None else '',
               'reinforced_l': f'{result_read_exel[4]:.3f}'.replace(".", ',') if result_read_exel[4] != None else '',
               'rubble_l': f'{result_read_exel[5]:.3f}'.replace(".", ',') if result_read_exel[5] != None else '',
               'crushed_stone_l': f'{result_read_exel[6]:.3f}'.replace(".", ',') if result_read_exel[6] != None else '',
               'ground_crushed_l': f'{result_read_exel[7]:.3f}'.replace(".", ',') if result_read_exel[7] != None else '',
               'priming_l': f'{result_read_exel[8]:.3f}'.replace(".", ',') if result_read_exel[8] != None else '',
               'asphalt_w': f'{result_read_exel[9]:.3f}'.replace(".", ',') if result_read_exel[9] != None else '',
               'reinforced_w': f'{result_read_exel[10]:.3f}'.replace(".", ',') if result_read_exel[10] != None else '',
               'rubble_w': f'{result_read_exel[11]:.3f}'.replace(".", ',') if result_read_exel[11] != None else '',
               'crushed_stone_w': f'{result_read_exel[12]:.3f}'.replace(".", ',') if result_read_exel[12] != None else '',
               'ground_crushed_w': f'{result_read_exel[13]:.3f}'.replace(".", ',') if result_read_exel[13] != None else '',
               'priming_w': f'{result_read_exel[14]:.3f}'.replace(".", ',') if result_read_exel[14] != None else ''}

    answer_lenght = f'{result_read_exel[2]:.3f}'.split(".")
    go_word['covering_weight1'] = answer_lenght[0]
    go_word['covering_weight2'] = answer_lenght[1]

    # lenght_all = list(map(lambda x: float(x) if x != None else 0, result_read_exel[18:22]))
    ans_acf = False
    for i in range(count_none + 1):
        weight_all1 = list(map(lambda x: float(x) if x != None else 0, [sheet1[f'J{k + i}'].value,
                                                                       sheet1[f'K{k + i}'].value,
                                                                       sheet1[f'L{k + i}'].value,
                                                                       sheet1[f'M{k + i}'].value,
                                                                       sheet1[f'N{k + i}'].value,
                                                                       sheet1[f'O{k + i}'].value]))
        if weight_all1[0] != 0:
            ans_acf = True
            break
    index_row = []
    for i in range(len(list(wb1))):
        wb1.active = i
        sheet = wb1.active
        if i == 0:
            sheet_one(count_n=count_none + 1, numb_k=k)
        elif i == 1:
            index_row.append(sheet_two(count_n=count_none + 1, numb_k=k))
        elif i == 2:
            index_row.append(sheet_three(count_n=count_none + 1, numb_k=k))
        elif i == 3:
            index_row.append(sheet_four(count_n=count_none + 1, numb_k=k))
        elif i == 4:
            index_row.append(sheet_five(count_n=count_none + 1, numb_k=k))
        elif i == 5:
            sheet_six(count_n=count_none + 1, numb_k=k)
        elif i == 6:
            sheet_seven(ans=ans_acf)

    if not os.path.isdir(go_word['name_road']):
        os.mkdir(go_word['name_road'])
    os.chdir(go_word['name_road'])
    wb1.save(f'отчет {go_word["name_road"]}.xlsx')

    wb1 = openpyxl.reader.excel.load_workbook(filename=f'отчет {go_word["name_road"]}.xlsx', data_only=True)
    cccc = 0
    for i in [1, 2, 3, 4]:
        wb1.active = i
        sheet = wb1.active
        print(result[cccc] == sheet[f'A{index_row[cccc]}'].value, sep='\n')
        cccc += 1
    wb1.close()
    # wb1.save(f'отчет {go_word["name_road"]}.xlsx'))
    print("After close")
    print(index_row)
    cccc = 0
    wb1 = openpyxl.reader.excel.load_workbook(filename=f'отчет {go_word["name_road"]}.xlsx', data_only=True)
    for i in [1, 2, 3, 4]:
        wb1.active = i
        sheet = wb1.active
        while (result[cccc] != sheet[f'A{index_row[cccc]}'].value):
            sheet[f'A{index_row[cccc]}'].value = result[cccc]
            wb1.save(f'отчет {go_word["name_road"]}.xlsx')
            wb1 = openpyxl.reader.excel.load_workbook(filename=f'отчет {go_word["name_road"]}.xlsx', data_only=True)
            print(False)
            baaaaad.add(go_word["name_road"])
        print(True)
        # print(sheet[f'A{index_row[cccc]}'].value, result[i], result[i] == sheet[f'A{index_row[cccc]}'].value, sep='\n')
        cccc += 1
    wb1.close()
    os.chdir("..")
    k += 1
    print(k - 10, result_read_exel[0])
    wb1.close()

wb.close()